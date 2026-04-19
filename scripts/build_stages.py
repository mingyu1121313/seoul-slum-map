#!/usr/bin/env python3
"""
build_stages.py
사업단계 정규화 + 진입일자 흡수 → data/stages.json
- xlsx (재개발/도시정비형 460건): 단계별 진입일자 풀 데이터
- layerA.json (모아타운/주환개선/가로주택/지역주택): 현재 단계만 정규화
- manual_overrides: 8개 핵심 사이트 + 해제 이력 수동 입력
출력: data/stages.json = { id: {bucket, group, dates, first_year, history, years_stalled} }
"""
import json, sys, re
from pathlib import Path
from datetime import date

ROOT = Path(__file__).parent.parent
DATA = ROOT / "data"
XLSX = ROOT / "★(25년 12월기준) 서울시 정비사업 추진현황.xlsx"
OUT  = DATA / "stages.json"
TODAY_YEAR = 2026

# ── 12개 버킷 (3그룹) ──
BUCKETS = {
    '추진준비':   '추진중',
    '연번부여':   '추진중',
    '안전진단':   '추진중',
    '대상지선정': '추진중',
    '구역지정':   '진행중',
    '추진위원회': '진행중',
    '조합설립인가':'진행중',
    '건축심의':   '진행중',
    '사업시행인가':'진행중',
    '관리처분인가':'진행중',
    '이주':       '진행중',
    '착공':       '완료',
    '준공':       '완료',
    '청산':       '완료',
}

# ── stage 문자열 → 버킷 매핑 ──
STAGE_NORMALIZE = {
    '조합설립인가':'조합설립인가', '조합창립총회':'조합설립인가', '조합규약작성':'조합설립인가',
    '관리지역고시':'구역지정', '관리지역지정(공람)':'추진준비',
    '추진위원회승인':'추진위원회', '추진위구성':'추진위원회',
    '정비계획 수립':'구역지정', '정비구역지정':'구역지정',
    '사업시행인가':'사업시행인가', '사업계획승인':'사업시행인가', '시공사선정':'사업시행인가',
    '대상지선정':'대상지선정',
    '관리처분인가':'관리처분인가',
    '조합원 모집신고':'추진위원회',
    '관리계획수립':'추진준비', '관리계획 수립':'추진준비',
    '착공':'착공', '분양':'착공', '철거 및 착공':'착공',
    '이전고시':'준공', '준공인가':'준공',
    '철거':'이주',
    '위원회심의':'건축심의', '지구단위계획수립/건축심의/교통심의':'건축심의', '통합심의':'건축심의', '건축심의':'건축심의',
    '수립범위자문':'추진준비', '사전자문':'추진준비',
}

# ── xlsx 컬럼 → 단계명 ──
XLSX_STAGES = [
    (11, '구역지정'),    # 최초
    (13, '추진위원회'),
    (14, '조합설립인가'),
    (15, '건축심의'),
    (16, '사업시행인가'),
    (18, '관리처분인가'),
    (20, '이주'),
    (22, '착공'),
]

# ── 8개 핵심 사이트 + 동급 케이스 수동 보정 ──
# id별로 첫 area 지정연도(area_first_year) + 해제이력 + 모아타운/주환개선 진입일
MANUAL = {
    # 현저동 1-5 (모아타운 moa1350): 같은 area가 2005년 주환개선 지정 → 20년 정체 → 2024 모아타운 신청
    'moa1350': {
        'bucket': '추진준비', 'first_year': 2024,
        'area_first_year': 2005,
        'dates': {'대상지선정': '2024-08', '관리계획수립': '2025-06-19'},
        'history': [
            {'event': '주환개선 지정', 'year': 2005, 'note': '동일 area, 합의 실패로 20년 방치'},
            {'event': '모아타운 신청', 'year': 2024},
            {'event': '관리계획 고시', 'year': 2025, 'date': '2025-06-19'},
        ],
    },
    # 현저2 (주환개선 jb03): 2005 지정, 정체
    'jb03': {
        'bucket': '구역지정', 'first_year': 2005,
        'area_first_year': 2005,
        'dates': {'구역지정': '2005'},
        'history': [{'event': '정비구역 지정', 'year': 2005}],
    },
    # 충정로1 (la0489): 2009 지정 → 2017 해제 → 2024.12 재지정
    'la0489': {
        'bucket': '조합설립인가', 'first_year': 2024,
        'area_first_year': 2009,
        'dates': {'구역지정': '2024-12-19', '조합설립인가': '2025-03-26'},
        'history': [
            {'event': '1차 정비구역 지정', 'year': 2009},
            {'event': '구역 해제', 'year': 2017, 'note': '사업 정체로 해제'},
            {'event': '재지정 (공공재개발)', 'year': 2024, 'date': '2024-12-19'},
            {'event': '조합설립인가', 'year': 2025, 'date': '2025-03-26'},
        ],
    },
    # 방배동 977 (moa1851): 2021 후보 → 2023.12 고시 → 조합설립인가
    'moa1851': {
        'bucket': '조합설립인가', 'first_year': 2021,
        'area_first_year': 2021,
        'dates': {'대상지선정': '2021-11', '관리지역고시': '2023-12-14', '조합설립인가': '2024'},
        'history': [
            {'event': '서초구 1호 모아타운 후보', 'year': 2021},
            {'event': '관리지역 지정고시', 'year': 2023, 'date': '2023-12-14'},
        ],
    },
    # 방배동 977,978,980 (moa2911): 977과 같은 후보지만 978/980 미고시
    'moa2911': {
        'bucket': '추진준비', 'first_year': 2021,
        'area_first_year': 2021,
        'dates': {'대상지선정': '2021-11'},
        'history': [
            {'event': '모아타운 후보 (977,978,980 통합)', 'year': 2021},
            {'event': '977 분리 고시 / 978·980 미고시', 'year': 2023},
        ],
    },
    # 원효로4가 30-1 (la0929): 2022 모아타운 후보 → 2024.11 조합설립
    'la0929': {
        'bucket': '조합설립인가', 'first_year': 2022,
        'area_first_year': 2022,
        'dates': {'대상지선정': '2022-10', '조합설립인가': '2024-11'},
        'history': [
            {'event': '모아타운 대상지 선정', 'year': 2022},
            {'event': '조합설립인가', 'year': 2024, 'date': '2024-11'},
        ],
    },
}

def normalize_bucket(stage_str):
    if not stage_str: return None
    s = stage_str.strip()
    return STAGE_NORMALIZE.get(s)

def cell_year(v):
    if v is None: return None
    if hasattr(v, 'year'): return v.year
    s = str(v).strip()
    m = re.search(r'(19|20)\d{2}', s)
    return int(m.group(0)) if m else None

def cell_date_str(v):
    if v is None: return None
    if hasattr(v, 'isoformat'): return v.isoformat()[:10]
    return str(v).strip()

def normalize_name(s):
    if not s: return ''
    s = s.replace(' ', '').replace('·','').replace('일대','').replace('일원','')
    # 긴 접미사 먼저 제거 (순서 중요)
    for kw in ['주택재개발정비사업조합','재개발정비사업조합','주택재개발정비사업',
               '재개발정비사업','주거환경개선사업','가로주택정비사업',
               '주택정비형재개발','공공재개발사업','공공재개발',
               '재정비촉진구역','재정비촉진지구','재정비촉진',
               '주택정비형','정비사업','정비구역',
               '주택재개발','주거환경개선','재개발','재건축',
               '구역','지구','조합','사업']:
        s = s.replace(kw, '')
    s = re.sub(r'제(\d)', r'\1', s)
    return s.strip()

def load_xlsx_stages():
    """xlsx → {정규화 이름: {dates, first_year, current_bucket}}"""
    from openpyxl import load_workbook
    wb = load_workbook(XLSX, data_only=True)
    ws = wb['Sheet1']
    out = {}
    for r in range(5, ws.max_row + 1):
        name = ws.cell(r, 3).value
        if not name: continue
        loc = ws.cell(r, 4).value or ''
        cur_stage = ws.cell(r, 9).value or ''
        # cur_stage 라벨 → 버킷 (xlsx는 단순 라벨: 구역지정/추진위/조합설립/건축심의/사업시행/관리처분/이주/착공/준공)
        bucket_map = {'구역지정':'구역지정','추진위':'추진위원회','추진위원회':'추진위원회',
                      '조합설립':'조합설립인가','건축심의':'건축심의',
                      '사업시행':'사업시행인가','관리처분':'관리처분인가',
                      '이주':'이주','착공':'착공','준공':'준공'}
        bucket = bucket_map.get(cur_stage.strip(), normalize_bucket(cur_stage))

        dates = {}
        years = []
        for col, label in XLSX_STAGES:
            v = ws.cell(r, col).value
            ds = cell_date_str(v)
            yr = cell_year(v)
            if ds: dates[label] = ds
            if yr: years.append(yr)
        first_year = min(years) if years else None
        key = normalize_name(str(name))
        out[key] = {
            'name_xlsx': str(name),
            'loc': str(loc),
            'bucket': bucket,
            'dates': dates,
            'first_year': first_year,
        }
    return out

def main():
    sys.stdout.reconfigure(encoding='utf-8')
    layerA = json.loads((DATA / 'layerA.json').read_text(encoding='utf-8'))
    items = layerA['items']

    xlsx_data = load_xlsx_stages() if XLSX.exists() else {}
    print(f'xlsx 매핑: {len(xlsx_data)}건')

    out = {}
    for it in items:
        iid = it.get('id')
        if not iid: continue
        # 1) MANUAL 우선
        if iid in MANUAL:
            m = MANUAL[iid].copy()
            m['group'] = BUCKETS.get(m['bucket'], '진행중')
            yr = m.get('area_first_year') or m.get('first_year')
            m['years_stalled'] = (TODAY_YEAR - yr) if yr else None
            out[iid] = m
            continue
        # 2) xlsx 매칭 시도 (이름 정규화)
        nname = normalize_name(it.get('name',''))
        xrow = xlsx_data.get(nname)
        if xrow and xrow['bucket']:
            bucket = xrow['bucket']
            yr = xrow['first_year']
            out[iid] = {
                'bucket': bucket,
                'group': BUCKETS.get(bucket, '진행중'),
                'dates': xrow['dates'],
                'first_year': yr,
                'years_stalled': (TODAY_YEAR - yr) if yr else None,
                'source': 'xlsx',
            }
            continue
        # 3) layerA의 stage 문자열 정규화
        bucket = normalize_bucket(it.get('stage',''))
        if bucket:
            out[iid] = {
                'bucket': bucket,
                'group': BUCKETS.get(bucket, '진행중'),
                'dates': {},
                'first_year': None,
                'years_stalled': None,
                'source': 'layerA',
            }
        else:
            out[iid] = {
                'bucket': None, 'group': None,
                'dates': {}, 'first_year': None, 'years_stalled': None,
                'source': 'unknown',
            }

    # layerB도 최소 필드만
    try:
        layerB = json.loads((DATA / 'layerB.json').read_text(encoding='utf-8'))
        for it in layerB.get('items', []):
            iid = it.get('id')
            if iid and iid not in out:
                out[iid] = {'bucket': None, 'group': None, 'dates': {}, 'first_year': None, 'years_stalled': None, 'source': 'layerB'}
    except Exception:
        pass

    OUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding='utf-8')
    # 통계
    bcount = {}
    for v in out.values():
        b = v['bucket'] or '(미분류)'
        bcount[b] = bcount.get(b, 0) + 1
    print(f'총 {len(out)}건 → {OUT}')
    for b, c in sorted(bcount.items(), key=lambda x:-x[1]):
        print(f'  {b}: {c}')

if __name__ == '__main__':
    main()
